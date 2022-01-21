import pandas
from nodes import Node, DictNodes, Runtime, withjinja
from sqlalchemy import create_engine
import pysftp
from ftplib import FTP
import tempfile
import os
from openpyxl import load_workbook
from webdav3.client import Client as ClientWebdav
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from oauth2client.service_account import ServiceAccountCredentials

__version__ = "0.9.0"


def to_fwf(df, fname="", names=[], colspecs=[]):
    ln = 0
    for col in colspecs:
        ln = max(ln, col[1])
    data = []
    for row in range(0, len(df.index)):
        string = " " * ln
        colidx = 0
        for name in names:
            if name in df.index.names:
                value = str(df.index[row][df.index.names.index(name)])
            else:
                value = str(df.iloc[row][name])
            size = colspecs[colidx][1] - colspecs[colidx][0] - 1
            value = value.ljust(size, ' ')[:size]
            string = string[0:colspecs[colidx][0]] + value + string[colspecs[colidx][0]+size:]
            colidx = colidx + 1
        data.append(string)
    content = "\n".join(data)
    if len(fname) > 0:
        open(fname, "w").write(content)
    return content


class LocalServer():

    def __init__(self, **kw):
        for key in kw:
            if isinstance(kw[key], str) is True:
                self.__setattr__(key, kw[key].strip())
            else:
                self.__setattr__(key, kw[key])
            self._cli = None

    def connect(self):
        pass

    def close(self):
        pass

    def put(self, filin, filout):
        pass

    def get(self, filout, filin):
        pass

    def clean(self, fil):
        pass


class WebdavServer(LocalServer):

    def connect(self):
        options = {'webdav_hostname': self.locserver,
                'webdav_login': self.locuser,
                'webdav_password': self.locpassword}
        self._cli = ClientWebdav(options)

    def close(self):
        pass

    def put(self, local_path, remote_path):
        self._cli.upload_sync(remote_path=remote_path, local_path=local_path)

    def get(self, local_path, remote_path):
        self._cli.download_sync(remote_path=remote_path, local_path=local_path)

    def clean(self, fil):
        self._cli.clean(fil)

    def isexist(self, fil):
        return self._cli.check(fil)


class Sharepoint365Server(LocalServer):

    def connect(self):
        ctx_auth = AuthenticationContext(self.locserver)
        ctx_auth.acquire_token_for_user(self.locuser, self.locpassword)
        self.locsite = self.locserver + '/' + self.locsite
        self._cli = ClientContext(self.locsite, ctx_auth)

    def close(self):
        pass

    def put(self, local_path, remote_path):
        remote_path = self.locsite + remote_path
        with open(local_path, 'rb') as content_file:
            file_content = content_file.read()
        self.debug(File.save_binary(self._cli, remote_path, file_content))

    def get(self, local_path, remote_path):
        remote_path = self.locsite + remote_path
        response = File.open_binary(self._cli, remote_path)
        with open(local_path, "wb") as local_file:
            local_file.write(response.content)

    def clean(self, fil):
        fil = self.locsite + fil
        file_to_delete = self._cli.web.get_file_by_server_relative_url(fil)
        file_to_delete.delete_object()
        self.debug(self._cli.execute_query())

    def isexist(self, fil):
        fil = self.locsite + fil
        try:
            self._cli.web.get_file_by_server_relative_url(fil)
            return True
        except Exception:
            return False


class FtpServer(LocalServer):

    def connect(self):
        self._cli = FTP()
        self._cli.connect(host=self.locserver, port=self.locport)
        self._cli.login(self.locuser, self.locpassword)

    def close(self):
        self._cli.close()

    def put(self, local_path, remote_path):
        with open(local_path, "rb") as file:
            self._cli.storbinary("STOR %s" % remote_path, file)

    def get(self, local_path, remote_path):
        with open(local_path, "wb") as file:
            self._cli.retrbinary("RETR %s" % remote_path, file.write)

    def clean(self, fil):
        self._cli.delete(fil)

    def isexist(self, fil):
        try:
            self._cli.size(fil)
            return True
        except Exception:
            return False


class SftpServer(LocalServer):

    def connect(self):
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None
        self._cli = pysftp.Connection(self.locserver, username=self.locuser, password=self.locpassword, port=self.locport, cnopts=cnopts)

    def close(self):
        self._cli.close()

    def put(self, local_path, remote_path):
        self._cli.put(local_path, remote_path)

    def get(self, local_path, remote_path):
        self._cli.get(remote_path, local_path)

    def clean(self, fil):
        self._cli.remove(fil)

    def isexist(self, fil):
        return self._cli.exists(fil)


class GoogleDriveServer(LocalServer):

    def _getid(self, paths, create_if_missing=False):
        id = 'root'
        for elt in [path for path in paths.split('/') if len(path) > 0]:
            folders = self._cli.ListFile({"q": "'%s' in parents and trashed=false" % id}).GetList()
            try:
                id = [folder for folder in folders if folder['title'] == elt][0]['id']
            except Exception:
                if create_if_missing is True:
                    file1 = self._cli.CreateFile({'title': elt, "mimeType": "application/vnd.google-apps.folder"})
                    file1.Upload()
                    id = file1['id']
                else:
                    raise Exception("%s not found" % paths)
        return id

    def connect(self):
        gauth = GoogleAuth()
        scope = ["https://www.googleapis.com/auth/drive"]
        dirpath = tempfile.mkdtemp()
        pathsecret = os.path.join(dirpath, 'secret.json')
        with open(pathsecret, 'w') as fil:
            fil.write(self.secret)
        gauth.credentials = ServiceAccountCredentials.from_json_keyfile_name(pathsecret, scope)
        self._cli = GoogleDrive(gauth)

    def close(self):
        pass

    def put(self, local_path, remote_path):
        fil = self._cli.CreateFile({'title': remote_path.split("/")[-1], 'parents': [{'id': self._getid('/'.join(remote_path.split('/')[:-1]), create_if_missing=True)}]})
        with open(local_path) as f:
            contents = f.read()
        fil.SetContentString(contents)
        fil.Upload()

    def get(self, local_path, remote_path):
        id = self._getid(remote_path)
        fil = self._cli.CreateFile({'id': id})
        fil.GetContentFile(filename=id)
        content_bytes = fil.content
        string_data = content_bytes.read().decode('utf-8')
        with open(local_path, 'w') as filout:
            filout.write(string_data)

    def clean(self, fil):
        fil = self._cli.CreateFile({'id': self._getid(fil)})
        fil.Delete()

    def isexist(self, fil):
        try:
            self._getid(fil)
            return True
        except Exception:
            return False


class RuntimeDb(Runtime):
    def __init__(self, id, idrun, name, log, **kw):
        Runtime.__init__(self, id, idrun, name, ['in1', ], log)
        if len(kw['port'].strip()) == 0:
            kw['port'] = "1433"
        for key in kw:
            if isinstance(kw[key], str) is True:
                self.__setattr__(key, kw[key].strip())
            else:
                self.__setattr__(key, kw[key])
        if self.method.strip() == "":
            self.method = None
        if len(self.optioncnx.strip()) > 0:
            self.optioncnx = "?%s" % self.optioncnx
        if self.dropindex == 'true':
            self.dropindex = False
        else:
            self.dropindex = True
        if len(self.locport) > 0:
            self.locport = int(self.locport)
        else:
            if self.loc == 'sftp':
                self.locport = 22

    def run(self, in1):
        table = withjinja(self.table).render({"in1": in1})
        query = withjinja(self.query).render({"in1": in1})
        pathfile = withjinja(self.pathfile).render({"in1": in1})
        sheet = withjinja(self.sheet).render({"in1": in1})
        colspecs = [(int(obj["start"]), int(obj["end"])) for obj in self.fields]
        names = [obj["field"] for obj in self.fields]
        if self.type in ('mssql', 'pgsql', 'mysql', 'sqlite'):
            self.debug("connect to //%s:********@%s:%s/%s%s" % (self.user, self.server, self.port, self.database, self.optioncnx))
            if self.type == 'mssql':
                cnx_str = "mssql+pymssql://%s:%s@%s:%s/%s%s" % (self.user, self.password, self.server, self.port, self.database, self.optioncnx)
                con = create_engine(cnx_str)
            if self.type == 'pgsql':
                cnx_str = "pgsql+psycopg2://%s:%s@%s:%s/%s%s" % (self.user, self.password, self.server, self.port, self.database, self.optioncnx)
                con = create_engine(cnx_str)
            if self.type == 'mysql':
                cnx_str = "pgsql+pymysql://%s:%s@%s:%s/%s%s" % (self.user, self.password, self.server, self.port, self.database, self.optioncnx)
            if self.type == 'sqlite':
                cnx_str = "sqlite:///%s" % self.url
                con = create_engine(cnx_str)
            if self.type == 'odbc':
                cnx_str = self.url
                con = create_engine(cnx_str)
            if self.action == 'read':
                if len(table) > 0:
                    query = "select * from %s" % table
                self.debug("read sql %s" % query)
                return pandas.read_sql_query(query, con=con)
            if self.action == 'write':
                self.debug("run sql")
                in1.to_sql(self.table, con=con, if_exists=self.ifexists, method=self.method, index=self.dropindex)
                return in1
        if self.type in ('csv', 'xlsx', 'json', 'xml', 'fwf'):
            if self.action == 'read':
                if self.loc == 'local':
                    self.debug("read local %s" % pathfile)
                    if self.type == 'csv':
                        in1 = pandas.read_csv(pathfile, sep=self.delimiter)
                    if self.type == 'xlsx':
                        in1 = pandas.read_excel(pathfile, sheet_name=sheet)
                    if self.type == 'json':
                        in1 = pandas.read_json(pathfile)
                    if self.type == 'xml':
                        in1 = pandas.read_xml(pathfile)
                    if self.type == 'fwf':
                        in1 = pandas.read_fwf(pathfile, skiprows=int(self.skiprows), skipfooter=int(self.skipfooter), colspecs=colspecs, names=names)
                    if self.delafter == 'true':
                        os.remove(pathfile)
                    return in1
                if self.loc in ('webdav', 'sftp', 'ftp', 'sharepoint365', 'googledrive'):
                    self.debug("connect %s %s" % (self.loc, self.locserver))
                    if self.loc == 'webdav':
                        cli = WebdavServer(locserver=self.locserver, locuser=self.locuser, locpassword=self.locpassword)
                    if self.loc == 'sftp':
                        cli = SftpServer(locserver=self.locserver, locuser=self.locuser, locpassword=self.locpassword, locport=self.locport)
                    if self.loc == 'ftp':
                        cli = FtpServer(locserver=self.locserver, locuser=self.locuser, locpassword=self.locpassword, locport=self.locport)
                    if self.loc == 'sharepoint365':
                        cli = Sharepoint365Server(locserver=self.locserver, locuser=self.locuser, locpassword=self.locpassword, locport=self.locport, locsite=self.locsite)
                    if self.loc == 'googledrive':
                        cli = GoogleDriveServer(secret=self.secret)
                    cli.connect()
                    dirpath = tempfile.mkdtemp()
                    path = os.path.join(dirpath, pathfile.split("/")[-1])
                    self.debug("download webdav %s" % self.pathfile)
                    cli.get(remote_path=pathfile, local_path=path)
                    cli.close()
                    self.debug("read local %s" % path)
                    if self.type == 'csv':
                        in1 = pandas.read_csv(path, sep=self.delimiter)
                    if self.type == 'xlsx':
                        in1 = pandas.read_excel(path, sheet_name=sheet)
                    if self.type == 'json':
                        in1 = pandas.read_json(path)
                    if self.type == 'xml':
                        in1 = pandas.read_xml(path)
                    if self.type == 'xml':
                        in1 = pandas.read_xml(path)
                    if self.delafter == 'true':
                        cli.clean(pathfile)
            if self.action == 'write':
                if self.loc == 'local':
                    if (os.path.isdir(os.path.dirname(os.path.abspath(pathfile)))) is False:
                        self.debug("local mkdir %s" % os.path.dirname(os.path.abspath(pathfile)))
                        os.makedirs(os.path.dirname(os.path.abspath(pathfile)))
                    self.debug("local write %s" % pathfile)
                    if self.type == 'csv':
                        in1.to_csv(pathfile, sep=self.delimiter, index=self.dropindex)
                    if self.type == 'xlsx':
                        if os.path.isfile(pathfile) is True:
                            book = load_workbook(pathfile)
                            writer = pandas.ExcelWriter(pathfile, engine='openpyxl')
                            writer.book = book
                            in1.to_excel(writer, sheet_name=sheet, index=self.dropindex)
                            writer.save()
                            writer.close()
                        else:
                            in1.to_excel(pathfile, sheet_name=sheet, index=self.dropindex)
                    if self.type == 'json':
                        in1.to_json(pathfile, index=self.dropindex)
                    if self.type == 'xml':
                        in1.to_xml(pathfile, index=self.dropindex)
                    if self.type == 'fwf':
                        to_fwf(in1, pathfile, names=names, colspecs=colspecs)
                if self.loc in ('webdav', 'sftp', 'ftp', 'sharepoint365', 'googledrive'):
                    self.debug("connect %s %s" % (self.loc, self.locserver))
                    if self.loc == 'webdav':
                        cli = WebdavServer(locserver=self.locserver, locuser=self.locuser, locpassword=self.locpassword)
                    if self.loc == 'sftp':
                        cli = SftpServer(locserver=self.locserver, locuser=self.locuser, locpassword=self.locpassword, locport=self.locport)
                    if self.loc == 'ftp':
                        cli = FtpServer(locserver=self.locserver, locuser=self.locuser, locpassword=self.locpassword, locport=self.locport)
                    if self.loc == 'sharepoint365':
                        cli = Sharepoint365Server(locserver=self.locserver, locuser=self.locuser, locpassword=self.locpassword, locport=self.locport, locsite=self.locsite)
                    if self.loc == 'googledrive':
                        cli = GoogleDriveServer(secret=self.secret)
                    cli.connect()
                    dirpath = tempfile.mkdtemp()
                    path = os.path.join(dirpath, pathfile.split("/")[-1])
                    self.debug("write loc %s" % path)
                    if self.type == 'csv':
                        in1.to_csv(path, sep=self.delimiter, index=self.dropindex)
                    if self.type == 'xlsx':
                        if cli.isexist(pathfile) is True:
                            cli.get(remote_path=pathfile, local_path=path)
                        if os.path.isfile(path) is True:
                            book = load_workbook(path)
                            writer = pandas.ExcelWriter(path, engine='openpyxl')
                            writer.book = book
                            in1.to_excel(writer, sheet_name=sheet, index=self.dropindex)
                            writer.save()
                            writer.close()
                        else:
                            in1.to_excel(path, sheet_name=sheet, index=self.dropindex)
                    if self.type == 'json':
                        in1.to_json(path, index=self.dropindex)
                    if self.type == 'xml':
                        in1.to_xml(path, index=self.dropindex)
                    if self.type == 'fwf':
                        to_fwf(in1, path, names=names, colspecs=colspecs)
                    if cli.isexist(pathfile) is True:
                        cli.clean(pathfile)
                    self.debug("upload webdav %s to %s" % (path, pathfile))
                    cli.put(remote_path=pathfile, local_path=path)
                    cli.close()
                return in1


class Db(Node):
    def __init__(self, *args, **kw):
        Node.__init__(self, *args, **kw)
        DictNodes()['db'] = RuntimeDb
